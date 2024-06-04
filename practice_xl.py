from openpyxl import Workbook, load_workbook

# wb = Workbook()

wb = load_workbook("example.xlsx")
# ws = wb.active
ws = wb.create_sheet(title = "Sheet2")
# ws["A1"] = 1
# ws["B1"] = 2
# ws["A2"] = 3
# ws["B2"] = 4
ws.cell(1, 1, 1)
ws.cell(1, 2, 2)
ws.cell(2, 1, 3)
ws.cell(2, 2, 4)

# ws.append([1,2,3,4])
# print(ws.cell(1,1).value)

# for row in ws.rows:
#     for cell in row:
#         print(cell.value)

for row in ws.iter_rows(min_row=1, min_col=1, max_row=8, max_col=10):
	for cell in row:
		print(cell.value)

for col in ws.iter_cols(min_row=1, min_col=1, max_row=10, max_col=10):
    for cell in col:
        print(cell.value)
		
wb.save("test.xlsx")
# print(ws)                 

