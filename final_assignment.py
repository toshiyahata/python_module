import pyodbc
import pandas as pd
from openpyxl import Workbook, load_workbook
import os 
from dotenv import load_dotenv
load_dotenv()

#**************************************
SERVER = os.getenv("SERVER")
USER = os.getenv("DB_USER")
PWD = os.getenv("DB_PASSWORD")
DATABASE = os.getenv("DATABASE")

sql = """CREATE TABLE movies (
        id int IDENTITY(1,1),
        Release_Date Date,
        Title varchar(300),
        Overview varchar(1000),
        Popularity float,
        Vote_Count int,
        Vote_Average float,
        Original_Language varchar(50),
        Genre varchar(100),
        Poster_Url varchar(300)
        )"""
#**************************************


conn_str = 'Driver={SQL Server};Server=' + SERVER + ';Uid=' + USER + ';pwd=' + PWD + ';Database=' + DATABASE + ';'
conn = pyodbc.connect(conn_str)
cursor = conn.cursor()


#　データ取り込みから挿入まで
"""
elements="Release_Date, Title, Overview, Popularity, \
        Vote_Count, Vote_Average, Original_Language,\
        Genre, Poster_Url"
    
cursor.execute(sql)

data = pd.read_csv("mymoviedb.csv", chunksize=50)
for chunk in data:
    chunk = chunk.dropna(how = "any")
    for index, row in chunk.iterrows():
        cursor.execute("INSERT INTO movies (" + elements + ") values(?,?,?,?,?,?,?,?,?)",\
                       row.Release_Date, row.Title, row.Overview, row.Popularity, row.Vote_Count,\
                          row.Vote_Average, row.Original_Language, row.Genre, row.Poster_Url)
    
conn.commit()
"""
df = pd.read_sql('SELECT * FROM movies', conn)
conn.close()

wb = Workbook()
ws1 = wb.active
ws1.title = "Language Counts"
ws2 = wb.create_sheet(title = "Genre Ratings")

#Language Counts
lc_list = df.groupby(["Original_Language"])["Original_Language"].count()

count = 1
ws1.cell(1,1,"国")
ws1.cell(1,2,"数")
for country, number in lc_list.items():
    count += 1
    ws1.cell(count, 1, country)
    ws1.cell(count, 2, number)

#Genre Ratings
gr_list = df.groupby(["Genre"])["Vote_Average"].mean()

count = 1
ws2.cell(1,1,"Genre")
ws2.cell(1,2,"Rating")
for genre, rating in gr_list.items():
    count += 1
    ws2.cell(count, 1, genre)
    ws2.cell(count, 2, rating)

wb.save("Final_八幡.xlsx")
