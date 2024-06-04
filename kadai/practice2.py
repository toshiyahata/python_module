from openpyxl import Workbook, load_workbook

wb = Workbook()
ws = wb.active
ws.title = "TestSheet"

#課題1
a=4
b=4
count = 1
for i in range(a):
    for j in range(b):
        ws.cell(i+2, j+2, count)
        count += 1

#課題2
sum = 0

for i in range(a):
    for j in range(b):
        sum += ws.cell(i+2, j+2).value

average = sum / (a*b)

ws.cell(8, 2, "平均")
ws.cell(9, 2, average)
ws.cell(8, 3, "合計")
ws.cell(9, 3, sum)

#追加課題
ws_ = wb.create_sheet(title = "DataInput")
for i in range(10):
    ws_.cell(i+1, 1, i+1)

for i in range(10):
    ws.cell(i+2, 6, ws_.cell(i+1,1).value+10)


wb.save("ExcelResult_八幡.xlsx")